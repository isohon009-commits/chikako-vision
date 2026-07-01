"""
FastAPI server — agent ilovasi rasm yuboradi, panel ma'lumot oladi.

Endpointlar:
  GET  /panel    — web boshqaruv paneli (brauzerda ochasiz)
  GET  /stats    — umumiy statistika (panel uchun)
  GET  /visits   — tashriflar ro'yxati (panel uchun)
  POST /analyze  — faqat AI sifat baholash
  POST /verify   — to'liq tekshiruv (rasm + GPS + dublikat) + bazaga saqlash
"""

import os
import time
import tempfile
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
from analyzer import analyze_image_bytes
from verify import verify_visit
from database import (get_visits, get_stats, get_stores, add_store, delete_store,
                      start_work, end_work, get_attendance, get_my_visits, get_agents)

app = FastAPI(title="Chikako Vision API")

HERE = os.path.dirname(__file__)
DATA_DIR = os.environ.get("DATA_DIR", HERE)
PHOTO_DIR = os.path.join(DATA_DIR, "photos")
os.makedirs(PHOTO_DIR, exist_ok=True)


@app.get("/")
def health():
    return {"status": "ishlayapti", "panel": "/panel"}


@app.get("/panel")
def panel():
    """Web boshqaruv paneli."""
    return FileResponse(os.path.join(HERE, "panel.html"))


@app.get("/agent")
def agent_app():
    """Agent ilovasi (telefonda ochiladi)."""
    return FileResponse(os.path.join(HERE, "agent.html"))


@app.get("/stores-page")
def stores_page():
    """Do'konlar bazasini boshqarish (admin)."""
    return FileResponse(os.path.join(HERE, "stores.html"))


# ---- Do'konlar API (A) ----
@app.get("/stores")
def stores_list():
    return get_stores()


@app.post("/stores")
async def stores_add(name: str = Form(...), lat: float = Form(...), lng: float = Form(...)):
    add_store(name, lat, lng)
    return {"ok": True}


@app.delete("/stores/{store_id}")
def stores_delete(store_id: int):
    delete_store(store_id)
    return {"ok": True}


@app.post("/stores/import")
async def stores_import(file: UploadFile = File(...)):
    """Excel'dan do'konlarni ommaviy yuklaydi (hamma eskisini almashtiradi)."""
    import io
    from openpyxl import load_workbook
    from database import replace_all_stores

    data = await file.read()
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel o'qib bo'lmadi: {e}")
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Fayl bo'sh")

    # Sarlavhadan ustunlarni topamiz
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    def find_col(*keys):
        for i, h in enumerate(header):
            if any(k in h for k in keys):
                return i
        return None
    name_i = find_col("назван", "nom", "name", "do'kon")
    gps_i = find_col("gps", "координат", "koordinat")

    if name_i is None or gps_i is None:
        raise HTTPException(status_code=400,
            detail="Kerakli ustunlar topilmadi (do'kon nomi va GPS koordinatasi)")

    out, skipped = [], 0
    for r in rows[1:]:
        try:
            name = r[name_i]
            gps = r[gps_i]
            if not name or not gps:
                skipped += 1
                continue
            parts = str(gps).replace(" ", "").split(",")
            lat, lng = float(parts[0]), float(parts[1])
            out.append((str(name).strip(), lat, lng))
        except Exception:
            skipped += 1

    if not out:
        raise HTTPException(status_code=400, detail="Hech qanday koordinatali do'kon topilmadi")

    replace_all_stores(out)
    return {"imported": len(out), "skipped": skipped}


# ---- Davomat API (C) ----
@app.post("/attendance/start")
async def attendance_start(agent: str = Form(...), date: str = Form(...)):
    t = start_work(agent, date)
    return {"start_time": t}


@app.post("/attendance/end")
async def attendance_end(agent: str = Form(...), date: str = Form(...)):
    t = end_work(agent, date)
    return {"end_time": t}


@app.get("/attendance")
def attendance_list(date: str):
    return get_attendance(date)


@app.get("/my-visits")
def my_visits(agent: str, date: str):
    return get_my_visits(agent, date)


@app.get("/stats")
def stats(date: str = None):
    return get_stats(date=date)


@app.get("/visits")
def visits(suspicious: bool = False, agent: str = None, store: str = None, date: str = None):
    return get_visits(only_suspicious=suspicious, agent=agent, store=store, date=date)


@app.get("/agents")
def agents_list():
    return get_agents()


@app.get("/photos/{filename}")
def get_photo(filename: str):
    """Saqlangan polka rasmini ko'rsatadi."""
    # xavfsizlik: faqat fayl nomi, yo'l emas
    safe = os.path.basename(filename)
    path = os.path.join(PHOTO_DIR, safe)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Rasm topilmadi")
    return FileResponse(path)


@app.post("/analyze")
async def analyze(file: UploadFile = File(...)):
    """Faqat AI sifat baholash."""
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Faqat rasm yuboring")
    image_bytes = await file.read()
    try:
        return analyze_image_bytes(image_bytes, file.content_type)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Tahlil xatosi: {e}")


@app.post("/verify")
async def verify(
    file: UploadFile = File(...),
    photo_lat: float = Form(None),
    photo_lng: float = Form(None),
    store_lat: float = Form(None),
    store_lng: float = Form(None),
    agent: str = Form(""),
    store: str = Form(""),
    date: str = Form(""),
):
    """To'liq tekshiruv: sifat + GPS + dublikat. Natija bazaga saqlanadi."""
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Faqat rasm yuboring")

    image_bytes = await file.read()
    suffix = ".png" if file.content_type == "image/png" else ".jpg"

    # Rasmni doimiy saqlaymiz (panelda ko'rsatish uchun)
    fname = f"{int(time.time()*1000)}{suffix}"
    fpath = os.path.join(PHOTO_DIR, fname)
    with open(fpath, "wb") as f:
        f.write(image_bytes)

    try:
        result = verify_visit(
            fpath,
            photo_lat=photo_lat, photo_lng=photo_lng,
            store_lat=store_lat, store_lng=store_lng,
            agent=agent, store=store, date=date, photo_file=fname,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Tekshiruv xatosi: {e}")

    return result
